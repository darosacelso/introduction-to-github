# Base de datos en Excel  
import openpyxl        
  
book = openpyxl.Workbook()       
Producto_page = book.create_sheet('Producto')  
Producto_page.append(['PRODUCTO', 'CANTIDAD', 'PRECIO'])     
  
productos = [  
    ['arroz', '5', '$120,00'],  
    ['Café', '3', '$80,00'],       
    ['azúcar', '2', '$50,00'],        
    ['sal', '1', '$30,00'],           
]    
  
for producto in productos:  
    Producto_page.append(producto)    
  
Cliente_page = book.create_sheet('Cliente')  
nombre_cliente = 'Celso'    
direccion_cliente = 'Calle Alberti 1568'   
dni_cliente = '9363XX74'  
valor_impuesto = '$33,00'        
  
Cliente_page.append(['NOMBRE DEL CLIENTE', 'DIRECCIÓN', 'DNI', 'VALOR DEL IMPUESTO'])      
Cliente_page.append([nombre_cliente, direccion_cliente, dni_cliente, valor_impuesto])      
  
def consultar_producto(nombre_producto):  
    for producto in productos:  
        if producto[0].lower() == nombre_producto.lower():  
            return f"Producto: {producto[0]}, Cantidad: {producto[1]}, Precio: {producto[2]}"  
    return "Producto no encontrado."  
  
resultado_consulta = consultar_producto("arroz")  
print(resultado_consulta)    
book.save('Productos Alimenticios.xlsx')

